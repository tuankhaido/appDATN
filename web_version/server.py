import os
import json
import pandas as pd
import joblib
import numpy as np
import io
import openpyxl
import logging
from flask import Flask, jsonify, request, send_from_directory, Response, make_response
from pathlib import Path
import pickle

app = Flask(__name__, static_folder='.')

# Cấu hình logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Đường dẫn tới file Excel
EXCEL_FILE_PATH = './DATN/appDATN/attached_assets/GiaoDien_KhungChuongTrinh.xlsx'

# Đường dẫn tới các mô hình
MODEL_PATHS = {
    'nam1': './DATN/appDATN/attached_assets/best_model_Year_1_SVM.pkl',
    'nam2': './DATN/appDATN/attached_assets/best_model_Year_2_SVM.pkl',
    'nam3': './DATN/appDATN/attached_assets/best_model_Year_3_Logistic_Regression.pkl',
    'nam4': './DATN/appDATN/attached_assets/best_model_Year_4_Logistic_Regression.pkl'
}

# Cache cho mô hình
_model_cache = {}

# Dữ liệu về các môn học và số tín chỉ
subjects_data = []

# Định nghĩa chuyển đổi từ điểm chữ sang điểm số (GPA)
LETTER_TO_NUMERIC = {
    'A': 4.0,
    'B': 3.0,
    'C': 2.0,
    'D': 1.0,
    'F': 0.0
}

# Hàm chuyển điểm số sang điểm chữ theo thang điểm
def numeric_to_letter(score):
    try:
        score = float(score)
        if score >= 8.5:
            return 'A'
        elif score >= 7.0:
            return 'B'
        elif score >= 5.5:
            return 'C'
        elif score >= 4.0:
            return 'D'
        else:
            return 'F'
    except:
        return 'F'

# Hàm tải mô hình machine learning với caching và validation
def load_model(year):
    global _model_cache
    try:
        if year in _model_cache:
            logger.debug("Lấy mô hình cho năm %s từ cache", year)
            return _model_cache[year]
        
        model_path = MODEL_PATHS.get(year)
        if not model_path:
            logger.error("Không có đường dẫn mô hình cho năm %s", year)
            return None
        
        model_path = Path(model_path)
        if not model_path.exists():
            logger.error("Không tìm thấy file mô hình tại %s cho năm %s", model_path, year)
            return None
        
        with open(model_path, 'rb') as f:
            model = joblib.load(f)
        
        # Kiểm tra mô hình
        required_attrs = ['predict']
        optional_attrs = ['feature_names_in_']
        for attr in required_attrs:
            if not hasattr(model, attr):
                logger.error("Mô hình cho năm %s không hợp lệ: thiếu thuộc tính %s", year, attr)
                return None
        
        for attr in optional_attrs:
            if hasattr(model, attr):
                logger.debug("Mô hình cho năm %s có %s: %s", year, attr, getattr(model, attr))
        
        # Lưu mô hình vào cache
        _model_cache[year] = model
        logger.info("Đã tải và lưu mô hình cho năm %s từ %s vào cache", year, model_path)
        return model
    
    except pickle.UnpicklingError as e:
        logger.error("Lỗi unpickle mô hình cho năm %s: %s", year, str(e))
        return None
    except Exception as e:
        logger.error("Lỗi không xác định khi tải mô hình cho năm %s: %s", year, str(e))
        return None

# Hàm đọc dữ liệu từ Excel và chuyển thành JSON
def read_excel_to_json():
    try:
        df = pd.read_excel(EXCEL_FILE_PATH)
        global subjects_data
        subjects_data = []
        for _, row in df.iterrows():
            if pd.notna(row['Mã Học Phần']) and pd.notna(row['Học Kỳ']):
                subjects_data.append({
                    'tenHocPhan': row['Tên Học Phần'],
                    'maHocPhan': row['Mã Học Phần'],
                    'soTinChi': int(row['Số Tín Chỉ']) if pd.notna(row['Số Tín Chỉ']) else 0,
                    'hocKy': int(row['Học Kỳ'])
                })
        os.makedirs('data', exist_ok=True)
        with open('data/subjects.json', 'w', encoding='utf-8') as f:
            json.dump(subjects_data, f, ensure_ascii=False, indent=2)
        logger.info("Đã đọc và lưu dữ liệu môn học vào subjects.json")
        return subjects_data
    except Exception as e:
        logger.error("Lỗi khi đọc file Excel: %s", str(e))
        return []

# Hàm tính điểm trung bình theo tín chỉ (dựa trên điểm chữ)
def calculate_weighted_average(scores):
    total_weighted_score = 0
    total_credits = 0
    
    for score in scores:
        subject_name = score['subjectName']
        letter_grade = score['score'].upper()
        
        if letter_grade not in LETTER_TO_NUMERIC:
            logger.warning("Điểm %s không hợp lệ cho môn %s, bỏ qua", letter_grade, subject_name)
            continue
        subject_score = LETTER_TO_NUMERIC[letter_grade]
        
        subject_info = next((s for s in subjects_data if s['tenHocPhan'] == subject_name), None)
        
        if subject_info:
            credits = subject_info['soTinChi']
            total_weighted_score += subject_score * credits
            total_credits += credits
        else:
            logger.warning("Không tìm thấy môn học %s trong dữ liệu", subject_name)
    
    if total_credits > 0:
        avg_score = total_weighted_score / total_credits
    else:
        avg_score = 0
        logger.warning("Không có tín chỉ hợp lệ để tính GPA")
    
    logger.info("GPA tính được: %.2f", avg_score)
    return avg_score

# Hàm chuyển GPA sang loại tốt nghiệp (dự phòng)
def gpa_to_graduation_type(avg_score):
    if avg_score >= 3.6:
        return "Xuất sắc"
    elif avg_score >= 3.2:
        return "Giỏi"
    elif avg_score >= 2.5:
        return "Khá"
    elif avg_score >= 2.0:
        return "Trung bình"
    else:
        return "Tốt nghiệp không đúng hạn"

# Hàm dự đoán kết quả tốt nghiệp
def predict_graduation(year, scores):
    try:
        # Tính GPA làm dự phòng
        avg_score = calculate_weighted_average(scores)
        fallback_grad_type = gpa_to_graduation_type(avg_score)
        logger.info("Kết quả GPA dự phòng: Loại %s (GPA: %.2f)", fallback_grad_type, avg_score)
        
        # Tải mô hình
        model = load_model(year)
        if not model:
            logger.warning("Không tải được mô hình cho năm %s, sử dụng kết quả GPA dự phòng", year)
            result_message = f"Bạn tốt nghiệp Loại {fallback_grad_type}" if fallback_grad_type != "Tốt nghiệp không đúng hạn" else "Bạn ra trường không đúng hạn"
            logger.info("Kết quả cuối cùng: %s (dựa trên GPA dự phòng)", result_message)
            return {
                'status': 'success',
                'prediction': 1 if avg_score >= 2.0 else 0,
                'message': result_message,
                'average_score': round(avg_score, 2)
            }
        
        # Chuẩn bị dữ liệu đầu vào cho mô hình
 # Chuẩn bị dữ liệu đầu vào cho model: dùng subjectName thay vì subjectCode
        subject_scores = {}
        for score in scores:
            subject_name = score['subjectName']
            letter_grade = score['score'].upper()
            if letter_grade in LETTER_TO_NUMERIC:
                subject_scores[subject_name] = letter_grade
            else:
                logger.warning("Điểm chữ %s không hợp lệ cho môn %s", letter_grade, subject_name)

        # Tạo DataFrame với tên môn học làm cột
        df = pd.DataFrame([subject_scores])
        
        # Dự đoán với mô hình
        try:
            if hasattr(model, 'feature_names_in_'):
                for feature in model.feature_names_in_:
                    if feature not in df.columns:
                        df[feature] = 'F'
                df = df[model.feature_names_in_]
            
            prediction = model.predict(df)[0]
            logger.info("Mô hình năm %s dự đoán: Loại %s", year, prediction)
            
            valid_grad_types = ["Xuất Sắc", "Giỏi", "Khá", "Trung Bình", "Ra trường không đúng hạn"]
            if prediction in valid_grad_types:
                result_message = f"Bạn sẽ tốt nghiệp Loại {prediction}" if prediction != "Ra trường không đúng hạn" else "Bạn sẽ ra trường không đúng hạn"
                prediction_value = 1 if prediction != "Ra trường không đúng hạn" else 0
                logger.info("Kết quả cuối cùng: %s (dựa trên mô hình)", result_message)
            else:
                logger.warning("Dự đoán không hợp lệ từ mô hình: %s, sử dụng GPA dự phòng: %s", prediction, fallback_grad_type)
                result_message = f"Bạn tốt nghiệp Loại {fallback_grad_type}" if fallback_grad_type != "Tốt nghiệp không đúng hạn" else "Bạn ra trường không đúng hạn"
                prediction_value = 1 if avg_score >= 2.0 else 0
                logger.info("Kết quả cuối cùng: %s (dựa trên GPA dự phòng)", result_message)
            
            return {
                'status': 'success',
                'prediction': prediction_value,
                'message': result_message,
                'average_score': round(avg_score, 2)
            }
        except Exception as e:
            logger.error("Lỗi khi dự đoán với mô hình năm %s: %s, sử dụng GPA dự phòng: %s", year, str(e), fallback_grad_type)
            result_message = f"Bạn tốt nghiệp Loại {fallback_grad_type}" if fallback_grad_type != "Tốt nghiệp không đúng hạn" else "Bạn ra trường không đúng hạn"
            logger.info("Kết quả cuối cùng: %s (dựa trên GPA dự phòng)", result_message)
            return {
                'status': 'success',
                'prediction': 1 if avg_score >= 2.0 else 0,
                'message': result_message,
                'average_score': round(avg_score, 2)
            }
            
    except Exception as e:
        logger.error("Lỗi không xác định trong predict_graduation: %s, sử dụng GPA dự phòng: %s", str(e), fallback_grad_type)
        result_message = f"Bạn tốt nghiệp Loại {fallback_grad_type}" if fallback_grad_type != "Tốt nghiệp không đúng hạn" else "Bạn ra trường không đúng hạn"
        logger.info("Kết quả cuối cùng: %s (dựa trên GPA dự phòng)", result_message)
        return {
            'status': 'success',
            'prediction': 1 if avg_score >= 2.0 else 0,
            'message': result_message,
            'average_score': round(avg_score, 2)
        }

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/css/<path:path>')
def send_css(path):
    return send_from_directory('css', path)

@app.route('/js/<path:path>')
def send_js(path):
    return send_from_directory('js', path)

@app.route('/data/<path:path>')
def send_data(path):
    return send_from_directory('data', path)

@app.route('/api/subjects', methods=['GET'])
def get_subjects():
    try:
        with open('data/subjects.json', 'r', encoding='utf-8') as f:
            subjects_data = json.load(f)
        logger.info("Đã gửi dữ liệu môn học từ subjects.json")
        return jsonify(subjects_data)
    except Exception as e:
        logger.error("Lỗi khi đọc subjects.json: %s", str(e))
        return jsonify({'error': str(e)}), 500

@app.route('/api/submit', methods=['POST'])
def submit_scores():
    try:
        data = request.get_json()
        scores = data.get('scores', [])
        year = data.get('year', 'nam1')
        
        # Chuyển đổi điểm số sang điểm chữ
        for score in scores:
            try:
                numeric_score = float(score['score'])
                if numeric_score < 0 or numeric_score > 10:
                    return jsonify({'status': 'error', 'message': f"Điểm {numeric_score} không hợp lệ. Điểm phải từ 0 đến 10."}), 400
                score['score'] = numeric_to_letter(numeric_score)
                logger.debug("Chuyển đổi điểm cho %s: %.2f -> %s", score['subjectCode'], numeric_score, score['score'])
            except (ValueError, TypeError):
                return jsonify({'status': 'error', 'message': f"Điểm {score['score']} không hợp lệ. Điểm phải là số."}), 400
        
        # Lưu dữ liệu
        os.makedirs('data', exist_ok=True)
        with open('data/submissions.json', 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info("Đã lưu dữ liệu vào submissions.json")
        
        # Thực hiện dự đoán
        prediction_result = predict_graduation(year, scores)
        
        return jsonify(prediction_result)
    except Exception as e:
        logger.error("Lỗi trong submit_scores: %s", str(e))
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    try:
        data = request.get_json()
        scores = data.get('scores', [])
        year = data.get('year', 'nam1')
        avg_score = calculate_weighted_average(scores)
        
        # Tạo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kết quả điểm"
        
        for col in range(1, 6):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Thêm tiêu đề
        ws['A1'] = "Kết quả điểm học tập"
        font = openpyxl.styles.Font(bold=True, size=14)
        ws['A1'].font = font
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Thêm header
        headers = ["Học kỳ", "Mã học phần", "Môn học", "Số tín chỉ", "Điểm"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell.fill = openpyxl.styles.PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'), 
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )

        # Sắp xếp điểm theo học kỳ
        sorted_scores = sorted(scores, key=lambda x: x['semester'])
        
        # Thêm dữ liệu
        for i, score in enumerate(sorted_scores, 1):
            row = i + 3
            values = [
                f"Học kỳ {score['semester']}", 
                score['subjectCode'], 
                score['subjectName'], 
                score['credits'], 
                score['score'].upper()
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'), 
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
                if col in [1, 4, 5]:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Thêm điểm trung bình
        last_row = len(sorted_scores) + 4
        avg_cell = ws.cell(row=last_row, column=4, value="Điểm trung bình:")
        avg_cell.font = openpyxl.styles.Font(bold=True)
        avg_cell.alignment = openpyxl.styles.Alignment(horizontal='right')
        
        score_cell = ws.cell(row=last_row, column=5, value=round(avg_score, 2))
        score_cell.font = openpyxl.styles.Font(bold=True)
        score_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # Lưu workbook vào BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Tạo response
        response = make_response(excel_file.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=ket-qua-diem-{year}.xlsx'
        
        logger.info("Đã xuất file Excel cho năm %s", year)
        return response
    except Exception as e:
        logger.error("Lỗi trong export_excel: %s", str(e))
        return jsonify({'status': 'error', 'message': str(e)}), 500

if __name__ == '__main__':
    os.makedirs('data', exist_ok=True)
    if not os.path.exists('data/submissions.json'):
        with open('data/submissions.json', 'w', encoding='utf-8') as f:
            json.dump([], f)
    read_excel_to_json()
    logger.info("Khởi tạo ứng dụng và tạo subjects.json")
    app.run(host='0.0.0.0', port=5000)